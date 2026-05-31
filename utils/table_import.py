from __future__ import annotations

import csv
import io
import json
import re
import unicodedata
from pathlib import Path
from typing import Any, Dict, List


TYPE_MAP = {
    "qcm": "single_choice",
    "choix_unique": "single_choice",
    "single_choice": "single_choice",
    "multiple_choice": "single_choice",
    "vrai_faux": "true_false",
    "true_false": "true_false",
    "question_courte": "short_answer",
    "short_answer": "short_answer",
    "rapprochement": "matching",
    "rapprochement_idees": "matching",
    "appariement": "matching",
    "matching": "matching",
}


def normalize_header(value: Any) -> str:
    text = "" if value is None else str(value).strip()
    text = unicodedata.normalize("NFKD", text)
    text = text.encode("ascii", "ignore").decode("ascii")
    text = re.sub(r"[^a-zA-Z0-9]+", "_", text)
    return text.strip("_").lower()


def clean_cell(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def get_value(row: Dict[str, Any], *names: str) -> str:
    for name in names:
        key = normalize_header(name)
        value = clean_cell(row.get(key, ""))
        if value:
            return value
    return ""


def parse_int(value: Any, default: int = 80) -> int:
    try:
        return int(float(str(value).strip()))
    except Exception:
        return default


def normalize_type(value: str) -> str:
    key = normalize_header(value)
    return TYPE_MAP.get(key, key or "single_choice")


def read_csv_rows(uploaded_file) -> List[Dict[str, Any]]:
    raw = uploaded_file.getvalue()

    try:
        text = raw.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = raw.decode("latin-1")

    first_line = text.splitlines()[0] if text.splitlines() else ""
    try:
        dialect = csv.Sniffer().sniff(text[:2048], delimiters=";,")
        delimiter = dialect.delimiter
    except Exception:
        delimiter = ";" if ";" in first_line else ","

    reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
    rows = []
    for row in reader:
        clean_row = {normalize_header(k): clean_cell(v) for k, v in row.items() if k is not None}
        rows.append(clean_row)
    return rows


def read_xlsx_rows(uploaded_file) -> List[Dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("Le module openpyxl est manquant. Lance : pip install openpyxl") from exc

    uploaded_file.seek(0)
    wb = load_workbook(uploaded_file, data_only=True)
    ws = wb.active

    values = list(ws.iter_rows(values_only=True))
    if not values:
        return []

    headers = [normalize_header(v) for v in values[0]]
    rows = []

    for line in values[1:]:
        if not any(clean_cell(v) for v in line):
            continue
        row = {}
        for header, value in zip(headers, line):
            if header:
                row[header] = clean_cell(value)
        rows.append(row)

    return rows


def parse_pairs(row: Dict[str, Any]) -> List[Dict[str, str]]:
    pairs = []

    pairs_json = get_value(row, "pairs_json", "paires_json")
    if pairs_json:
        try:
            parsed = json.loads(pairs_json)
            if isinstance(parsed, list):
                for item in parsed:
                    if isinstance(item, dict):
                        left = clean_cell(item.get("left", ""))
                        right = clean_cell(item.get("right", ""))
                        if left or right:
                            pairs.append({"left": left, "right": right})
            if pairs:
                return pairs
        except Exception:
            pass

    for i in range(1, 7):
        left = get_value(
            row,
            f"pair_{i}_left",
            f"paire_{i}_gauche",
            f"left_{i}",
            f"gauche_{i}",
        )
        right = get_value(
            row,
            f"pair_{i}_right",
            f"paire_{i}_droite",
            f"right_{i}",
            f"droite_{i}",
        )
        if left or right:
            pairs.append({"left": left, "right": right})

    return pairs


def parse_feedbacks(row: Dict[str, Any]) -> Dict[str, str]:
    feedbacks = {}
    for letter in ["a", "b", "c", "d", "e", "f"]:
        feedback = get_value(
            row,
            f"feedback_{letter}",
            f"explication_{letter}",
            f"retour_{letter}",
        )
        if feedback:
            feedbacks[letter.upper()] = feedback

    true_feedback = get_value(row, "feedback_vrai", "explication_vrai")
    false_feedback = get_value(row, "feedback_faux", "explication_faux")
    if true_feedback:
        feedbacks["VRAI"] = true_feedback
    if false_feedback:
        feedbacks["FAUX"] = false_feedback

    return feedbacks


def build_quality_summary(questions: List[Dict[str, Any]]) -> Dict[str, Any]:
    if not questions:
        return {
            "average_score": 0,
            "low_quality_count": 0,
            "warning": "Aucune question importee.",
        }

    scores = [parse_int(q.get("quality_score", 80), 80) for q in questions]
    average = round(sum(scores) / len(scores), 1)
    low = len([s for s in scores if s < 70])

    return {
        "average_score": average,
        "low_quality_count": low,
        "warning": "Synthese calculee a partir du fichier importe.",
    }


def rows_to_quiz(rows: List[Dict[str, Any]], filename: str = "quiz") -> Dict[str, Any]:
    if not rows:
        raise ValueError("Le fichier ne contient aucune ligne exploitable.")

    title = ""
    for row in rows:
        title = get_value(row, "quiz_title", "titre_quiz", "titre")
        if title:
            break

    if not title:
        title = f"Quiz importe - {Path(filename).stem}"

    questions = []

    for row in rows:
        question_text = get_value(row, "question", "consigne", "enonce")
        if not question_text:
            continue

        qtype = normalize_type(get_value(row, "type", "question_type", "type_question"))

        options = []
        for letter in ["a", "b", "c", "d", "e", "f"]:
            option = get_value(row, f"option_{letter}", f"reponse_{letter}")
            if option:
                options.append(option)

        if qtype == "true_false" and not options:
            options = ["Vrai", "Faux"]

        pairs = parse_pairs(row)
        feedbacks = parse_feedbacks(row)

        correct_answer = get_value(
            row,
            "correct_answer",
            "reponse_correcte",
            "bonne_reponse",
            "correction",
        )

        if qtype == "matching" and not correct_answer:
            correct_answer = "Voir les paires de correction."

        accepted_answers_raw = get_value(
            row,
            "accepted_answers",
            "reponses_acceptees",
            "synonymes",
        )
        accepted_answers = [
            item.strip()
            for item in re.split(r"[|;]", accepted_answers_raw)
            if item.strip()
        ]

        question = {
            "type": qtype,
            "domain": get_value(row, "domain", "domaine") or "Non classe",
            "subdomain": get_value(row, "subdomain", "sous_domaine"),
            "learning_objective": get_value(row, "learning_objective", "objectif_pedagogique"),
            "difficulty": get_value(row, "difficulty", "niveau") or "Import",
            "cognitive_level": get_value(row, "cognitive_level", "niveau_cognitif"),
            "competency": get_value(row, "competency", "competence", "competence_evaluee"),
            "concept_evaluated": get_value(row, "concept_evaluated", "notion", "notion_evaluee"),
            "question": question_text,
            "options": options,
            "pairs": pairs,
            "correct_answer": correct_answer,
            "accepted_answers": accepted_answers,
            "feedbacks": feedbacks,
            "explanation": get_value(row, "explanation", "explication", "justification"),
            "source_excerpt": get_value(row, "source_excerpt", "extrait_source"),
            "remediation": get_value(row, "remediation", "piste_de_revision", "pistes_de_travail"),
            "quality_score": parse_int(get_value(row, "quality_score", "score_qualite"), 80),
            "quality_notes": get_value(row, "quality_notes", "notes_qualite") or "Question importee depuis un fichier structure.",
        }

        questions.append(question)

    if not questions:
        raise ValueError("Aucune question n'a ete trouvee. Verifie la colonne 'question'.")

    return {
        "quiz_title": title,
        "analysis_summary": "Quiz importe depuis un fichier CSV ou Excel structure.",
        "questions": questions,
        "quality_summary": build_quality_summary(questions),
        "note": "Quiz charge depuis un fichier. La mise en forme, l'historique, la passation et les restitutions sont disponibles.",
    }


def read_quiz_table(uploaded_file) -> Dict[str, Any]:
    name = uploaded_file.name.lower()

    if name.endswith(".csv"):
        rows = read_csv_rows(uploaded_file)
    elif name.endswith(".xlsx"):
        rows = read_xlsx_rows(uploaded_file)
    else:
        raise ValueError("Format non supporte. Utilise CSV ou XLSX.")

    return rows_to_quiz(rows, uploaded_file.name)


def template_csv_bytes() -> bytes:
    content = """quiz_title;module;domain;subdomain;type;difficulty;cognitive_level;learning_objective;competency;concept_evaluated;question;option_a;option_b;option_c;option_d;correct_answer;accepted_answers;explanation;feedback_a;feedback_b;feedback_c;feedback_d;remediation;source_excerpt;pair_1_left;pair_1_right;pair_2_left;pair_2_right;pair_3_left;pair_3_right
Examen civique naturalisation - Entrainement 1;Examen civique;Symboles de la Republique;Devise;single_choice;Debutant;Memorisation;Identifier les symboles de la Republique;Connaitre la devise republicaine;Devise;Quelle est la devise de la Republique francaise ?;Liberte, ordre, patrie;Liberte, egalite, fraternite;Travail, famille, patrie;Unite, force, justice;B;;La devise de la Republique francaise est Liberte, egalite, fraternite.;Incorrect. Cette formule n'est pas la devise officielle actuelle.;Correct. Liberte, egalite, fraternite est la devise officielle de la Republique francaise.;Incorrect. Cette devise est associee au regime de Vichy, pas a la Republique actuelle.;Incorrect. Cette formulation n'est pas la devise officielle.;Revoir les symboles de la Republique : devise, drapeau, hymne, Marianne.;La devise officielle de la Republique est Liberte, egalite, fraternite.;;;;;;
Examen civique naturalisation - Entrainement 1;Examen civique;Institutions;Parlement;single_choice;Intermediaire;Comprehension;Distinguer les institutions et leurs roles;Identifier l'institution qui vote la loi;Parlement;Qui vote les lois en France ?;Le Parlement;Le President de la Republique;Le Conseil constitutionnel;Le Gouvernement;A;;Le Parlement vote les lois. Il comprend l'Assemblee nationale et le Senat.;Correct. Le Parlement est l'institution qui vote les lois.;Incorrect. Le President promulgue les lois et nomme le Premier ministre, mais il ne vote pas les lois.;Incorrect. Le Conseil constitutionnel controle la conformite des lois a la Constitution.;Incorrect. Le Gouvernement conduit la politique nationale et peut proposer des projets de loi, mais il ne vote pas les lois.;Revoir le role du Parlement, du Gouvernement et du President.;Le Parlement comprend l'Assemblee nationale et le Senat.;;;;;;
Examen civique naturalisation - Entrainement 1;Examen civique;Symboles de la Republique;Symboles;matching;Debutant;Comprehension;Associer les symboles republicains a leur signification;Reconnaitre les symboles republicains;Symboles;Associez chaque symbole a sa signification.;;;;;Voir les paires;;Chaque symbole renvoie a une notion republicaine precise.;;;;;Revoir les principaux symboles de la Republique francaise.;Marianne represente la Republique, la Marseillaise est l'hymne national, bleu blanc rouge correspond au drapeau.;Marianne;Republique;La Marseillaise;Hymne national;Bleu, blanc, rouge;Drapeau francais
"""
    return content.encode("utf-8-sig")
